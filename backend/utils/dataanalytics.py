import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class Dataanalytics:
    COLUMN_MAP = {
        "student_name": "Student",
        "username": "Username",
        "classroom_name": "Classroom",
        "department_name": "Department",
        "attendance_date": "Date",
        "morning_hours": "Morning (h)",
        "afternoon_hours": "Afternoon (h)",
        "total_hours": "Total (h)",
        "expected_morning_hours": "Expected Morning (h)",
        "expected_afternoon_hours": "Expected Afternoon (h)",
        "expected_total_hours": "Expected Total (h)",
        "morning_attendance_rate": "Morning Rate (%)",
        "afternoon_attendance_rate": "Afternoon Rate (%)",
        "total_attendance_rate": "Total Rate (%)",
        "total_late_minutes": "Late (min)",
    }

    @staticmethod
    def get_excel(data: list[dict], file_path: str) -> str:
        """Build the presence report workbook and return the path written."""
        if not data:
            raise ValueError("No attendance rows to export")

        df = pd.DataFrame(data)
        cols = [c for c in Dataanalytics.COLUMN_MAP if c in df.columns]
        df = df[cols].rename(columns=Dataanalytics.COLUMN_MAP)
        df = df.sort_values(["Student", "Date"]).reset_index(drop=True)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Daily Presence", index=False)
            # placeholder sheet; formulas are written below once we know row counts
            df[["Student"]].drop_duplicates().to_excel(
                writer, sheet_name="Summary", index=False
            )

        Dataanalytics._add_summary_formulas(file_path, df)
        Dataanalytics._style(file_path)
        return file_path

    @staticmethod
    def _add_summary_formulas(file_path: str, df: pd.DataFrame):
        wb = load_workbook(file_path)
        daily = wb["Daily Presence"]
        summary = wb["Summary"]

        last_daily_row = daily.max_row  # header is row 1
        student_col = get_column_letter(df.columns.get_loc("Student") + 1)
        total_col = get_column_letter(df.columns.get_loc("Total (h)") + 1)
        expected_col = get_column_letter(df.columns.get_loc("Expected Total (h)") + 1)
        rate_col = get_column_letter(df.columns.get_loc("Total Rate (%)") + 1)
        late_col = get_column_letter(df.columns.get_loc("Late (min)") + 1)

        summary.cell(row=1, column=2, value="Total Hours")
        summary.cell(row=1, column=3, value="Expected Hours")
        summary.cell(row=1, column=4, value="Avg Rate (%)")
        summary.cell(row=1, column=5, value="Total Late (min)")

        for r in range(2, summary.max_row + 1):
            name_ref = f"A{r}"
            summary.cell(row=r, column=2).value = (
                f"=SUMIF('Daily Presence'!{student_col}2:{student_col}{last_daily_row},"
                f"{name_ref},'Daily Presence'!{total_col}2:{total_col}{last_daily_row})"
            )
            summary.cell(row=r, column=3).value = (
                f"=SUMIF('Daily Presence'!{student_col}2:{student_col}{last_daily_row},"
                f"{name_ref},'Daily Presence'!{expected_col}2:{expected_col}{last_daily_row})"
            )
            summary.cell(row=r, column=4).value = (
                f"=ROUND(AVERAGEIF('Daily Presence'!{student_col}2:{student_col}{last_daily_row},"
                f"{name_ref},'Daily Presence'!{rate_col}2:{rate_col}{last_daily_row}),2)"
            )
            summary.cell(row=r, column=5).value = (
                f"=SUMIF('Daily Presence'!{student_col}2:{student_col}{last_daily_row},"
                f"{name_ref},'Daily Presence'!{late_col}2:{late_col}{last_daily_row})"
            )

        wb.save(file_path)

    @staticmethod
    def _style(file_path: str):
        wb = load_workbook(file_path)
        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for col_cells in ws.columns:
                length = max(
                    (len(str(c.value)) for c in col_cells if c.value is not None),
                    default=0,
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 10), 40
                )

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial")

        wb.save(file_path)